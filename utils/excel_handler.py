"""
Excel İşlem Modülü
Import/Export ve formatlama işlemleri
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Tuple, Optional
import os


class ExcelHandler:
    """Excel dosya işlemlerini yöneten sınıf"""

    @staticmethod
    def import_excel(file_path: str, sheet_name: Optional[str] = None) -> Tuple[bool, any]:
        """
        Excel dosyasını oku
        Returns: (başarılı_mı, DataFrame veya hata_mesajı)
        """
        try:
            if not os.path.exists(file_path):
                return False, "Dosya bulunamadı!"

            # Sayfa ismi belirtilmediyse ilk sayfayı oku
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            return True, df

        except Exception as e:
            return False, f"Excel okuma hatası: {str(e)}"

    @staticmethod
    def get_sheet_names(file_path: str) -> Tuple[bool, any]:
        """
        Excel dosyasındaki sayfa isimlerini getir
        Returns: (başarılı_mı, sayfa_listesi veya hata_mesajı)
        """
        try:
            xl_file = pd.ExcelFile(file_path)
            return True, xl_file.sheet_names

        except Exception as e:
            return False, f"Sayfa isimleri okunamadı: {str(e)}"

    @staticmethod
    def export_to_excel(data: List[List], columns: List[str], file_path: str,
                        sheet_name: str = "Veri", styled: bool = True) -> Tuple[bool, str]:
        """
        Veriyi Excel'e aktar
        """
        try:
            # DataFrame oluştur
            df = pd.DataFrame(data, columns=columns)

            if styled:
                # Styled Excel export
                return ExcelHandler._export_styled(df, file_path, sheet_name)
            else:
                # Basit export
                df.to_excel(file_path, sheet_name=sheet_name, index=False)
                return True, f"Excel'e aktarıldı: {file_path}"

        except Exception as e:
            return False, f"Excel'e aktarma hatası: {str(e)}"

    @staticmethod
    def _export_styled(df: pd.DataFrame, file_path: str, sheet_name: str) -> Tuple[bool, str]:
        """Stillendirilmiş Excel export"""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Header stilleri
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center')

                # Border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Header formatla
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Sütun genişliklerini ayarla
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min((max_length + 2) * 1.2, 50)  # Max 50
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Veri satırlarını formatla (zebra pattern)
                for row_num in range(2, len(df) + 2):
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')

                        # Zebra pattern
                        if row_num % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            return True, f"Stillendirilmiş Excel oluşturuldu: {os.path.basename(file_path)}"

        except Exception as e:
            return False, f"Styled export hatası: {str(e)}"

    @staticmethod
    def export_multiple_tables(tables_data: Dict[str, pd.DataFrame],
                               file_path: str) -> Tuple[bool, str]:
        """
        Birden fazla tabloyu ayrı sayfalarda Excel'e aktar
        tables_data: {sheet_name: DataFrame}
        """
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in tables_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            return True, f"{len(tables_data)} tablo Excel'e aktarıldı"

        except Exception as e:
            return False, f"Çoklu export hatası: {str(e)}"

    @staticmethod
    def validate_excel_file(file_path: str) -> Tuple[bool, str]:
        """Excel dosyasını doğrula"""
        if not os.path.exists(file_path):
            return False, "Dosya bulunamadı!"

        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False, "Geçersiz Excel dosyası!"

        try:
            # Dosyayı açmayı dene
            pd.read_excel(file_path, nrows=1)
            return True, "Geçerli Excel dosyası"
        except Exception as e:
            return False, f"Excel dosyası hatalı: {str(e)}"

    @staticmethod
    def get_excel_info(file_path: str) -> Dict:
        """Excel dosyası hakkında bilgi al"""
        try:
            xl_file = pd.ExcelFile(file_path)

            info = {
                'file_name': os.path.basename(file_path),
                'file_size': os.path.getsize(file_path),
                'sheet_count': len(xl_file.sheet_names),
                'sheet_names': xl_file.sheet_names,
                'sheets_info': {}
            }

            # Her sayfa için bilgi
            for sheet in xl_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet)
                info['sheets_info'][sheet] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns)
                }

            return info

        except Exception as e:
            return {'error': str(e)}

    @staticmethod
    def dataframe_to_sql_insert(df: pd.DataFrame, table_name: str) -> List[str]:
        """DataFrame'i SQL INSERT komutlarına çevir"""
        insert_queries = []

        columns = ', '.join([f'`{col}`' for col in df.columns])

        for _, row in df.iterrows():
            values = []
            for val in row:
                if pd.isna(val):
                    values.append('NULL')
                elif isinstance(val, str):
                    # String escape
                    val_escaped = val.replace("'", "''")
                    values.append(f"'{val_escaped}'")
                else:
                    values.append(str(val))

            values_str = ', '.join(values)
            query = f"INSERT INTO `{table_name}` ({columns}) VALUES ({values_str});"
            insert_queries.append(query)

        return insert_queries